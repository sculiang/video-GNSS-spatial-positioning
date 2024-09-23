import numpy as np
import openpyxl
import matplotlib.pyplot as plt

noiseName='Band_pass+WN_Y'

# 参数设置
size = 200000  # 信号长度
center_freq = 5  # 中心频率（Hz）
bandwidth = 1  # 带宽（Hz）
sample_rate = 200000  # 采样率（Hz）
wn_mean = 0.0 #白噪声均值
wn_std = 6.27 #白噪声均方差

workbook_write=openpyxl.Workbook()
sheet_noise=workbook_write.create_sheet(noiseName)

def generate_band_pass_noise(size, center_freq, bandwidth, sample_rate, wn_mean, wn_std):
    t = np.arange(size) / sample_rate
    noise = np.random.normal(wn_mean, wn_std, size)
    carrier = np.cos(2 * np.pi * center_freq * t)
    window = np.hanning(size)
    bp_signal = noise * window * carrier
    return bp_signal


band_pass_WN_noise = generate_band_pass_noise(size, center_freq, bandwidth, sample_rate, wn_mean, wn_std)

for i in range(len(band_pass_WN_noise)):
    sheet_noise.cell(row=i+1,column=1).value=band_pass_WN_noise[i]/0.3048

workbook_write.save("./Non-Gaussian/"+noiseName+".xlsx")

# 可视化或保存噪声数据
plt.plot(band_pass_WN_noise)
plt.show()