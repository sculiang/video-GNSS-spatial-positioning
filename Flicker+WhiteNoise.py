import numpy as np
import matplotlib.pyplot as plt
import openpyxl

noiseName='Flicker+WN_Y'

# 参数设置
length = 200000  # 信号长度
fs = 200000  # 采样频率
wn_mean = 0.0 #白噪声均值
wn_std = 6.27 #白噪声均方差

workbook_write=openpyxl.Workbook()
sheet_noise=workbook_write.create_sheet(noiseName)

t = np.arange(length) / fs  # 时间向量

# 生成一个白噪声
white_noise = np.random.normal(wn_mean, wn_std, length)
print("signal:",white_noise)

# 定义一个函数来生成闪烁噪声
def create_flicker_noise(white_noise, power=1.0, seed=None):
    rng = np.random.default_rng(seed)
    flicker = rng.uniform(low=-power, high=power, size=len(white_noise))
    return white_noise + flicker

# 添加闪烁噪声
flicker_wn_noise = create_flicker_noise(white_noise, power=1.0)

for i in range(len(flicker_wn_noise)):
    sheet_noise.cell(row=i+1,column=1).value=flicker_wn_noise[i]/0.3048

workbook_write.save("./Non-Gaussian/"+noiseName+".xlsx")

# 绘制信号
plt.figure()
plt.plot(t, white_noise, label='Original Signal')
plt.plot(t, flicker_wn_noise, label='Noisy Signal with Flicker Noise')
plt.legend()
plt.show()

