import numpy as np
import openpyxl
import matplotlib.pyplot as plt

noiseName='Matern+WN_Y'

# 参数设置
length = 200000  # 信号长度
length_scale = 10  # 长度尺度
wn_mean = 0.0 #白噪声均值
wn_std = 6.27 #白噪声均方差
signal_variance = 6.27  # 信号方差
noise_variance = 6.27  # 噪声方差

workbook_write=openpyxl.Workbook()
sheet_noise=workbook_write.create_sheet(noiseName)

def generate_matern_noise(n, length_scale, signal_variance, noise_variance):
    t = np.linspace(0, length_scale, n)
    signal = np.sqrt(signal_variance) * np.cos(2 * np.pi / length_scale * t)
    noise_std = np.sqrt(noise_variance)
    noise = np.random.normal(0, noise_std, n)
    return signal + noise

# 生成Matérn噪声
matern_WN_noise = generate_matern_noise(length, length_scale, signal_variance, noise_variance)

for i in range(len(matern_WN_noise)):
    sheet_noise.cell(row=i+1,column=1).value=matern_WN_noise[i]/0.3048

workbook_write.save("./Non-Gaussian/"+noiseName+".xlsx")

# 可视化或保存噪声数据
plt.plot(matern_WN_noise)
plt.show()