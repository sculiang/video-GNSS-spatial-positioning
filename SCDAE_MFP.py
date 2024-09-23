import torch
from torch import optim
import torch.nn.functional as F
import torch.nn as nn
import matplotlib.pyplot as plt
from data.dataset import MyDataSet, MyDataSet_Autoencoder
from data.load_datasets import x_train, y_train, z_train, x_test, y_test, z_test, fileName
import os
import numpy as np
import openpyxl
import shutil
from parameters_setting import batch_size, d_input, sequence_length
from torch.utils.data import DataLoader
from self_attention_operation.batch_error_calculate import BatchErrorCalculate

saveFileName=fileName

device=torch.device("cuda" if torch.cuda.is_available() else "cpu")

workbook_write=openpyxl.Workbook()
workbook_write_loss_scale=openpyxl.Workbook()
worksheet_write_x=workbook_write.create_sheet('final data x')
worksheet_write_z=workbook_write.create_sheet('final data z')
worksheet_write_loss_scale=workbook_write_loss_scale.create_sheet('loss scale')

args_save_dir = './SCDAE/'+fileName+'/checkPoint'
args_resume='./SCDAE/'+fileName+'/checkPoint/checkpoint.pth'
best_args_resume='./SCDAE/'+fileName+'/checkPoint/model_best.pth'
args_result_dir='./SCDAE/'+fileName+'/Result'
epoches=200
learning_rate=0.002
args_test_every=1

"""数据集导入"""
dataset_train = MyDataSet(x_train, y_train, z_train)
loader_train = DataLoader(dataset_train, batch_size, shuffle=True, drop_last=True)
dataset_test = MyDataSet(x_test, y_test, z_test)
loader_test = DataLoader(dataset_test, batch_size, shuffle=False, drop_last=True)
torch.set_default_tensor_type(torch.DoubleTensor)

def loss_function(x_hat,x,z_hat,z):
    loss_mse_x = F.mse_loss(x_hat,x,reduction='mean')
    loss_mse_z = F.mse_loss(z_hat, z, reduction='mean')
    loss_scale=loss_mse_z/loss_mse_x
    loss=loss_mse_x + 10*loss_mse_z
    return loss,loss_scale

def loss_test_in_training(z_hat,z):
    return F.mse_loss(z_hat, z, reduction='mean')

def save_checkpoint(state, is_best, outdir):
    """
    每训练一定的epochs后， 判断损失函数是否是目前最优的，并保存模型的参数
    :param state: 需要保存的参数，数据类型为dict
    :param is_best: 说明是否为目前最优的
    :param outdir: 保存文件夹
    :return:
    """
    if not os.path.exists(outdir):
        os.makedirs(outdir)

    checkpoint_file = os.path.join(outdir, 'checkpoint.pth')  # join函数创建子文件夹，也就是把第二个参数对应的文件保存在'outdir'里
    best_file = os.path.join(outdir, 'model_best.pth')
    torch.save(state, checkpoint_file)  # 把state保存在checkpoint_file文件夹中
    if is_best:
        shutil.copyfile(checkpoint_file, best_file)

def test_in_training(model,optimizer,loader_test,epoch,best_test_loss):
    test_avg_loss=0
    with torch.no_grad():#这一部分不计算梯度，也就是不放入计算图中去
        for encoder_input, gt, denoised_encoder_input in loader_test:
            x_input_test, z_gt, denoised_current_x_test = encoder_input.to(device), gt.to(device), denoised_encoder_input.to(device)
            # print("z_gt:",z_gt)
            z_gt = z_gt[:, sequence_length - 1, :]
            x_input_test = torch.reshape(x_input_test, (batch_size, d_input * sequence_length))
            x_hat_test, z_hat_test = model(x_input_test)
            loss_test = loss_test_in_training(z_hat_test,z_gt)  # 计算损失值，即目标函数
            test_avg_loss += loss_test
        batch_num = len(x_test)/batch_size
        test_avg_loss = test_avg_loss / batch_num
        is_best = test_avg_loss < best_test_loss
        best_test_loss = min(test_avg_loss,best_test_loss)
        save_checkpoint({
            'epoch':epoch,
            'best_test_loss':best_test_loss,
            'state_dict':model.state_dict(),
            'optimizer':optimizer.state_dict(),
        },is_best,args_save_dir)

        return best_test_loss

def main():
    model=Autoencoder().to(device)
    best_test_loss = np.finfo('f').max
    optimizer = optim.Adam(model.parameters(), lr=learning_rate)

    if os.path.isfile(best_args_resume):
        print('=>loading checkpoint %s' % best_args_resume)
        checkpoint = torch.load(best_args_resume)
        # start_epoch=checkpoint['epoch']+1
        best_test_loss = checkpoint['best_test_loss']
        model.load_state_dict(checkpoint['state_dict'])
        optimizer.load_state_dict(checkpoint['optimizer'])
        print('=>loaded checkpoint %s' % best_args_resume)
    else:
        print('=>no checkpoint found at %s' % best_args_resume)

    loss_epoch = []

    for epoch_i in range(epoches):
        loss_epoches = []
        loss_scale_sum = 0
        index = 0
        for encoder_input, imm_ukf, denoised_encoder_input in loader_train:
            X, z, recon_X = encoder_input.to(device), imm_ukf.to(device), denoised_encoder_input.to(device)

            X = torch.reshape(X, (batch_size, d_input * sequence_length))
            recon_X = torch.reshape(recon_X, (batch_size, d_input * sequence_length))

            X_hat, z_hat = model(X)

            z = z[:, sequence_length - 1, :]
                                                   
            batch_loss, loss_scale_iter = loss_function(X_hat, recon_X, z_hat, z)

            optimizer.zero_grad()
            batch_loss.backward()
            optimizer.step()
            loss_epoches.append(batch_loss.item())
            loss_scale_sum += loss_scale_iter

            '计算batch_error'
            batch_error, batch_single_error = BatchErrorCalculate(z_hat.cpu().detach().numpy(),
                                                                  z.cpu().detach().numpy())
            index = index + 1

        """显示loss和error"""
        str_epoch = 'Epoch: ' + str(epoch_i + 1) + ' / ' + str(epoches)
        print(str_epoch, 'loss =', '{:.8f}'.format(batch_loss),
              'error =', '{:.8f}'.format(batch_error),
              'single error =', '{:.8f}'.format(batch_single_error))

        loss_scale = loss_scale_sum / len(x_train) * batch_size
        # print('loss_scale:{}'.format(loss_scale))
        if (epoch_i + 1) % args_test_every == 0:  # 每过args_test_every个EPOCH测试一次模型
            best_test_loss = test_in_training(model, optimizer, loader_test, epoch_i, best_test_loss)

        loss_epoch.append(np.sum(loss_epoches) / len(x_train) * batch_size)


    loss_test_average = 0
    excel_row_index = 0
    for encoder_input, gt, denoised_encoder_input in loader_test:
        x_input_test, z_gt, denoised_current_x_test = encoder_input.to(device), gt.to(device), denoised_encoder_input.to(device)
        z_gt = z_gt[:, sequence_length-1, :]
        # print("z_test:", z_test)
        x_input_test = torch.reshape(x_input_test, (batch_size, d_input * sequence_length))
        x_hat_test, z_hat_test = model(x_input_test)

        var_pre = loss_scale * F.mse_loss(x_hat_test, x_input_test, reduction='mean')
        z_fused_long = (z_gt[:, 0] * var_pre + z_hat_test[:, 0] * 55167) / (var_pre + 55167)
        z_fused_lat = (z_gt[:, 1] * var_pre + z_hat_test[:, 1] * 43902) / (var_pre + 43902)

        for j in range(0, batch_size):
            worksheet_write_z.cell(row=excel_row_index * batch_size + j + 1, column=1).value = z_hat_test[j, 0].item()
            worksheet_write_z.cell(row=excel_row_index * batch_size + j + 1, column=2).value = z_hat_test[j, 1].item()
            worksheet_write_z.cell(row=excel_row_index * batch_size + j + 1, column=3).value = z_fused_long[j].item()
            worksheet_write_z.cell(row=excel_row_index * batch_size + j + 1, column=4).value = z_fused_lat[j].item()
            worksheet_write_z.cell(row=excel_row_index * batch_size + j + 1, column=5).value = z_gt[j, 0].item()
            worksheet_write_z.cell(row=excel_row_index * batch_size + j + 1, column=6).value = z_gt[j, 1].item()
            # print("test data:{}/{}:".format(excel_row_index*batch_size + j +1,len(test_dataset)))
        excel_row_index = excel_row_index + 1
        print("test data:{}/{} batches:".format(excel_row_index, len(x_test) // batch_size))
        loss_test, loss_test_scale = loss_function(x_hat_test, x_input_test, z_hat_test, z_gt)
        loss_test_average += loss_test
    loss_test_average = loss_test_average / len(x_test) * batch_size

    workbook_write.save('./results/SCDAE/Non-Gaussian/results '+saveFileName+'0920.xlsx')
    return loss_epoch, loss_test_average


class Autoencoder(nn.Module):
    def __init__(self,input_dim=d_input*sequence_length,h_dim1=6,z_dim1=2):
        super(Autoencoder,self).__init__()
        # ###三层神经网络
        # #dimensions
        # self.input_dim = input_dim
        # self.h_dim1 = h_dim1
        # self.z_dim1 = z_dim1
        # #encoder
        # self.fc1=nn.Linear(input_dim,h_dim1)
        # self.fc2=nn.Linear(h_dim1,z_dim1)
        # #decoder
        # self.fc3=nn.Linear(z_dim1,h_dim1)
        # self.fc4=nn.Linear(h_dim1,input_dim)

        ###四层神经网络
        #dimensions
        self.input_dim = input_dim
        self.h_dim1 = h_dim1
        self.z_dim1 = z_dim1
        #encoder
        self.fc1=nn.Linear(input_dim,h_dim1)
        self.fc2 = nn.Linear(h_dim1, h_dim1)
        self.fc3=nn.Linear(h_dim1,z_dim1)
        #decoder
        self.fc4=nn.Linear(z_dim1,h_dim1)
        self.fc5 = nn.Linear(h_dim1, h_dim1)
        self.fc6=nn.Linear(h_dim1,input_dim)

    def forward(self,x):
        # ### 三层神经网络
        # h1=F.leaky_relu(self.fc1(x),negative_slope=0.001)
        # z=self.fc2(h1)
        # h2=F.leaky_relu(self.fc3(z),negative_slope=0.001)
        # x_hat=self.fc4(h2)
        # return x_hat,z

        ## 四层神经网络
        denoised_hat_current_x=F.leaky_relu(self.fc1(x),negative_slope=0.001)
        h2 = F.leaky_relu(self.fc2(denoised_hat_current_x), negative_slope=0.001)
        # z=self.fc3(h2)
        z = F.leaky_relu(self.fc3(h2), negative_slope=0.001)
        h3=F.leaky_relu(self.fc4(z),negative_slope=0.001)
        recon_current_x = F.leaky_relu(self.fc5(h3), negative_slope=0.001)
        # x_hat=self.fc6(recon_current_x)
        x_hat = F.leaky_relu(self.fc6(recon_current_x),negative_slope=0.001)
        return x_hat,z

if __name__=='__main__':
    loss_epoch, loss_average = main()
    print('loss_average:{}'.format(loss_average))
    plt.plot(loss_epoch)
    plt.xlabel('epoch')
    plt.ylabel('loss')
    plt.show()


