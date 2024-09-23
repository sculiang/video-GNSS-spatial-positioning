import openpyxl
import pandas as pd

print("The loading starts...")
datasets_workbook=openpyxl.load_workbook("./citySim/with IMMUKF and visual noises/excel/McCulloch@Seminole-10.xlsx") #对应定位任务的数据
gt_sheet=datasets_workbook.get_sheet_by_name('McCulloch@Seminole-01')


gt_rows=[]
for i in range(1,gt_sheet.max_row):
    if i%1000 == 0:
        print("Loading data:{}/{}".format(i,gt_sheet.max_row))
    row=[]
    for j in range(0, gt_sheet.max_column):
       row.append(gt_sheet.cell(row=i+1, column=j+1).value)
    gt_rows.append(row)