import openpyxl
import numpy as np
import torch

fileName="RoundaboutB1"

print("The loading starts...")
datasets_workbook=openpyxl.load_workbook('data/citySim/final datasets/SCDAE/'+fileName+'.xlsx')
# datasets_workbook=openpyxl.load_workbook('data/citySim/final datasets/SCDAE/Non-Gaussian/without Gaussian visual noises/'+fileName+'.xlsx')

print("The loading ends.")
xtrain_sheet=datasets_workbook.get_sheet_by_name('x_train')
ytrain_sheet=datasets_workbook.get_sheet_by_name('y_train')
ztrain_sheet=datasets_workbook.get_sheet_by_name('z_train')
xtest_sheet=datasets_workbook.get_sheet_by_name('x_test')
ytest_sheet=datasets_workbook.get_sheet_by_name('y_test')
ztest_sheet=datasets_workbook.get_sheet_by_name('z_test')

x_seq_len, y_seq_len, z_seq_len = 3, 3, 3
x_samp_len, y_samp_len, z_samp_len = 6, 2, 6

x_train = []
for i in range (0,xtrain_sheet.max_row):
    x_samp_seq = []
    for j in range (x_seq_len):
        x_samp = []
        for h in range (x_samp_len):
            samp_pnt = xtrain_sheet.cell(row=i+1, column= 1 + j*x_samp_len + h).value #样本点
            x_samp.append(samp_pnt) #形成[,,,,,,]样本
        x_samp_seq.append(x_samp) # 形成[[],[],[]]样本
    x_train.append(x_samp_seq) #形成[[[],[],[]],[[],[],[]]]样本
print("x_train is loaded")

y_train = []
for i in range (0,ytrain_sheet.max_row):
    y_samp_seq = []
    for j in range (y_seq_len):
        y_samp = []
        for h in range (y_samp_len):
            samp_pnt = ytrain_sheet.cell(row=i+1, column= 1 + j*y_samp_len + h).value #样本点
            y_samp.append(samp_pnt) #形成[,,,,,,]样本
        y_samp_seq.append(y_samp) # 形成[[],[],[]]样本
    y_train.append(y_samp_seq) #形成[[[],[],[]],[[],[],[]]]样本
print("y_train is loaded")

z_train = []
for i in range (0,ztrain_sheet.max_row):
    z_samp_seq = []
    for j in range (z_seq_len):
        z_samp = []
        for h in range (z_samp_len):
            samp_pnt = ztrain_sheet.cell(row=i+1, column= 1 + j*z_samp_len + h).value #样本点
            z_samp.append(samp_pnt) #形成[,,,,,,]样本
        z_samp_seq.append(z_samp) # 形成[[],[],[]]样本
    z_train.append(z_samp_seq) #形成[[[],[],[]],[[],[],[]]]样本
print("z_train is loaded")

x_test = []
for i in range(0, xtest_sheet.max_row):
    x_samp_seq = []
    for j in range(x_seq_len):
        x_samp = []
        for h in range(x_samp_len):
            samp_pnt = xtest_sheet.cell(row=i + 1, column=1 + j * x_samp_len + h).value  # 样本点
            x_samp.append(samp_pnt)  # 形成[,,,,,,]样本
        x_samp_seq.append(x_samp)  # 形成[[],[],[]]样本
    x_test.append(x_samp_seq)  # 形成[[[],[],[]],[[],[],[]]]样本
print("x_test is loaded")

y_test = []
for i in range(0, ytest_sheet.max_row):
    y_samp_seq = []
    for j in range(y_seq_len):
        y_samp = []
        for h in range(y_samp_len):
            samp_pnt = ytest_sheet.cell(row=i + 1, column=1 + j * y_samp_len + h).value  # 样本点
            y_samp.append(samp_pnt)  # 形成[,,,,,,]样本
        y_samp_seq.append(y_samp)  # 形成[[],[],[]]样本
    y_test.append(y_samp_seq)  # 形成[[[],[],[]],[[],[],[]]]样本
print("y_test is loaded")

z_test = []
for i in range(0, ztest_sheet.max_row):
    z_samp_seq = []
    for j in range(z_seq_len):
        z_samp = []
        for h in range(z_samp_len):
            samp_pnt = ztest_sheet.cell(row=i + 1, column=1 + j * z_samp_len + h).value  # 样本点
            z_samp.append(samp_pnt)  # 形成[,,,,,,]样本
        z_samp_seq.append(z_samp)  # 形成[[],[],[]]样本
    z_test.append(z_samp_seq)  # 形成[[[],[],[]],[[],[],[]]]样本
print("z_test is loaded")

np_x_train=np.array(x_train)
np_y_train=np.array(y_train)
np_z_train=np.array(z_train)
np_x_test=np.array(x_test)
np_y_test=np.array(y_test)
np_z_test=np.array(z_test)

x_train=torch.from_numpy(np_x_train)
y_train=torch.from_numpy(np_y_train)
z_train=torch.from_numpy(np_z_train)
x_test=torch.from_numpy(np_x_test)
y_test=torch.from_numpy(np_y_test)
z_test=torch.from_numpy(np_z_test)
