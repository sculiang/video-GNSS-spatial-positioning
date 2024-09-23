from data.make_dataset import x_train, y_train, z_train, x_test, y_test, z_test
import openpyxl
import numpy as np

workbook_write=openpyxl.Workbook()
sheet_xtrain=workbook_write.create_sheet('x_train')
sheet_ytrain=workbook_write.create_sheet('y_train')
sheet_ztrain=workbook_write.create_sheet('z_train')
sheet_xtest=workbook_write.create_sheet('x_test')
sheet_ytest=workbook_write.create_sheet('y_test')
sheet_ztest=workbook_write.create_sheet('z_test')

x_train=np.array(x_train)
y_train=np.array(y_train)
z_train=np.array(z_train)
x_test=np.array(x_test)
y_test=np.array(y_test)
z_test=np.array(z_test)

'''
x_train.shape[0] #训练样本个数
x_train.shape[1] #单样本序列长度,比如3
x_train.shape[2] #单样本长度,为6
'''

for i in range(x_train.shape[0]):
    for j in range(x_train.shape[1]):
        for h in range(x_train.shape[2]):
            sheet_xtrain.cell(row=1 + i,column=1 + j*x_train.shape[2] + h).value=x_train[i][j][h]
print("x_train is saved")

for i in range(y_train.shape[0]):
    for j in range(y_train.shape[1]):
        for h in range(y_train.shape[2]):
            sheet_ytrain.cell(row=1 + i,column=1 + j*y_train.shape[2] + h).value=y_train[i][j][h]
print("y_train is saved")

for i in range(z_train.shape[0]):
    for j in range(z_train.shape[1]):
        for h in range(z_train.shape[2]):
            sheet_ztrain.cell(row=1 + i,column=1 + j*z_train.shape[2] + h).value=z_train[i][j][h]
print("z_train is saved")

for i in range(x_test.shape[0]):
    for j in range(x_test.shape[1]):
        for h in range(x_test.shape[2]):
            sheet_xtest.cell(row=1 + i,column=1 + j*x_test.shape[2] + h).value=x_test[i][j][h]
print("x_test is saved")

for i in range(y_test.shape[0]):
    for j in range(y_test.shape[1]):
        for h in range(y_test.shape[2]):
            sheet_ytest.cell(row=1 + i,column=1 + j*y_test.shape[2] + h).value=y_test[i][j][h]
print("y_test is saved")

for i in range(z_test.shape[0]):
    for j in range(z_test.shape[1]):
        for h in range(z_test.shape[2]):
            sheet_ztest.cell(row=1 + i,column=1 + j*z_test.shape[2] + h).value=z_test[i][j][h]
print("z_test is saved")

workbook_write.save("citySim/final datasets/RoundaboutB1.xlsx")
# workbook_write.save("citySim/final datasets/Non-Gaussian/RandomWalk_WN/RoundaboutB.xlsx")
# workbook_write.save("UE4_ZUUU/final datasets/datasets for UE4_ZUUU to He Ruliang.xlsx")
